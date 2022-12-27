import openpyxl
import datetime


def write_to_excel(file1, file2, date):
    # Open Excel file
    workbook = openpyxl.load_workbook(file1)
    workbook2 = openpyxl.load_workbook(file2)

    # Select worksheet
    ws = workbook.worksheets[0]
    ws2 = workbook2.worksheets[0]

    # Get cell style
    my_yellow = openpyxl.styles.colors.Color(rgb='FFFF00')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
    my_style = openpyxl.styles.NamedStyle(name='highlight', fill=my_fill)

    # Get date of first day of current month
    finalDate = datetime.datetime.strptime(date, "%b%Y").strftime("%Y-%m-%d")

    # Get column index
    for row in ws.iter_rows():
        for cell in row:          
            if str(finalDate) in str(cell.value):
                index = cell.column - 2
                indexMonth = cell.column - 1
            elif cell.value == 'Ref.':
                indexRef = cell.column - 1

    # Create dictionary with data from previous month's Excel file
    num = 1
    key = 'Forecast ' + str(num)
    forecast = {key: []}

    for row in ws2.iter_rows():
        if row[indexRef].value == 'Forecast':
            for cell in row:
                if cell.column > indexRef and cell.value != None and cell.value != 0:
                    key = 'Forecast ' + str(num)
                    if key not in forecast:
                        forecast[key] = []
                    forecast[key].append(cell.value)
            num += 1

    size = len(forecast)

    # Replace keys with numbers in dictionary
    numForecast = {}
    for i, key in enumerate(forecast):
        numForecast[i] = forecast[key] 
                   
    # Loop to highlight cells
    num2 = 0
    num3 = 2
    for row in ws.iter_rows():
        if row[indexRef].value == 'Baseline 1' and row[index].value != row[indexMonth].value: # Check if Baseline or Actual
            row[indexMonth].style = my_style
        elif row[indexRef].value == 'Forecast':
            forecastList = numForecast[num2]                      
            for cell in row:               
                if num3 <= len(forecastList) - 1 and cell.column > indexMonth and cell.value != None and cell.value != 0 and cell.value != forecastList[num3]:                               
                    cell.style = my_style   
                    num3 += 1
        num2 += 1
        num3 = 2
        if num2 > size:
            break
    
    return ws
