import streamlit as st
import eap_roadmap
import eap_compare
import pandas as pd
import base64
import components.login as authenticator
import openpyxl


st.set_page_config(page_title="EAP Mero 2", page_icon=":house:", layout="wide") 
with open("css\main.css" ) as f:
   st.markdown(f"<style>{f.read()}</style>",unsafe_allow_html = True)

  
name, authentication_status, username = authenticator.run_authenticator()

if "visibility" not in st.session_state:
    st.session_state.visibility = "visible"
    st.session_state.disabled = False

if st.session_state["authenticated"] == None:
    st.warning("Please enter your username and password")

if st.session_state["authenticated"] == True: 
   st.sidebar.image("./Figures/technip_logo.png", use_column_width=True)
   st.sidebar.title(f"Welcome {name}")   
   
  
   uploaded_file = st.file_uploader('Choose a Mero 2 EAP file', type='xlsx')
 
   uploaded_file2 = st.file_uploader('Choose a Mero 2 EAP file with a month prior to the one chosen above', type='xlsx')
      
 
   
   if uploaded_file is not None:     
       #st.header("Sum Results")
       # Input Excel
       st.session_state.resultados = eap_roadmap.roadmap(uploaded_file) 
              
       
       df = pd.DataFrame(st.session_state.resultados.items(), columns=['LEVEL', 'SUM']) 
       df[['SUM', 'STATUS']] = df['SUM'].str.split(", ", 1, expand=True)
       df[['LEVEL', 'DESCRIPTION']] = df['LEVEL'].str.split(' -- ', 1, expand=True)
       
       
       if uploaded_file2 is None:
          option = st.sidebar.selectbox(
                 "Filter table",
                 (" ", "CORRECT SUM", "INCORRECT SUM")
             )
                
          
          optionMonth = st.sidebar.selectbox(
                 "Choose EAP Month",
                 (" ", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov","Dec"),
                 label_visibility=st.session_state.visibility,
                 disabled=st.session_state.disabled
             )
          
          
          optionYear = st.sidebar.selectbox(
                 "Choose EAP year",
                 (" ", "2020", "2021", "2022", "2023", "2024"),
                 label_visibility=st.session_state.visibility,
                 disabled=st.session_state.disabled
             )
          
         
          
          finalDate = optionMonth + optionYear   
          
          if option == " ":       
              st.dataframe(df.iloc[:,[0,3,1,2]].style.applymap(eap_roadmap.color_background, subset=['STATUS']))
          elif option == "INCORRECT SUM":
              df.query("STATUS == 'INCORRECT SUM'", inplace = True)
              st.dataframe(df.iloc[:,[0,3,1,2]].style.applymap(eap_roadmap.color_background, subset=['STATUS']))  
          else: 
              df.query("STATUS == 'CORRECT SUM'", inplace = True)
              st.dataframe(df.iloc[:,[0,3,1,2]].style.applymap(eap_roadmap.color_background, subset=['STATUS']))
              
   
       elif uploaded_file2 is not None:
           print("Upload realizado arquivo 2")
           
           optionMonth = st.sidebar.selectbox(
                  "Choose EAP Month",
                  (" ", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov","Dec"),
                  label_visibility=st.session_state.visibility,
                  disabled=st.session_state.disabled
              )
           
           
           optionYear = st.sidebar.selectbox(
                  "Choose EAP year",
                  (" ", "2020", "2021", "2022", "2023", "2024"),
                  label_visibility=st.session_state.visibility,
                  disabled=st.session_state.disabled
              )
           
          
           
           finalDate = optionMonth + optionYear
           
           check = st.sidebar.checkbox("Date selected", key="disabled")
           
           if check:                                                 
              #st.sidebar.markdown("Click [here](uploaded_file) to download the file.")
              #st.sidebar.download_button(label="Click here to download the file.", data=compareFiles, file_name='myresults.xlsx', mime='text/xlsx') 
              
              st.markdown("Click the button below to download the manipulated file")              
           
              if st.button("Download"):                

                  ws = eap_compare.write_to_excel(uploaded_file, uploaded_file2, finalDate) 
                  my_yellow = openpyxl.styles.colors.Color(rgb='FFFF00')
                  my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)            
                  
                  #workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
                  workbook = openpyxl.load_workbook(uploaded_file, data_only=False)
                  ws2 = workbook['EAP Mero 2'] 
                  workbook.remove(ws2)
                  new_ws = workbook.create_sheet('EAP Mero 2', 0)
                  
                  for row in ws.rows:
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)  # Copy cell value
                        new_cell.font = cell.font.copy()  # Copy font formatting
                        new_cell.border = cell.border.copy()  # Copy border formatting
                        new_cell.fill = cell.fill.copy()  # Copy fill formatting
                        new_cell.number_format = cell.number_format  # Copy number formatting
                        new_cell.protection = cell.protection.copy()  # Copy protection formatting
                        new_cell.alignment = cell.alignment.copy()  # Copy alignment formatting
                        
                  fileName = uploaded_file.name
                  workbook.save(filename=fileName)
                  with open(fileName, 'rb') as f:
                      excel_file = f.read()  
                  excel_file_b64 = base64.b64encode(excel_file).decode('utf-8')                                            
                  st.markdown(f'<a href="data:application/octet-stream;base64,{excel_file_b64}" download={fileName}>Download manipulated file</a>', unsafe_allow_html=True)                               
       
   else:
       st.warning('you need to upload an excel file.')
       
       optionMonth = st.sidebar.selectbox(
              "Choose EAP Month",
              ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov","Dec"),
              disabled = True,
              label_visibility = "hidden"
          )
       
       optionYear = st.sidebar.selectbox(
              "Choose EAP year",
              ("20", "21", "22", "23", "24"),
              disabled = True,
              label_visibility = "hidden"
          )
       
       add_selectbox = st.sidebar.selectbox(
            "Filter table",
            (" ","CORRECT SUM", "INCORRECT SUM"),
            disabled = True,
            label_visibility = "hidden"
        )
       
   authenticator.button_logout()
  
       
   

 
 

    



